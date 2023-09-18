from bs4 import BeautifulSoup
import requests
import openpyxl as xl

def get_price(barcode):
    url = "https://www.google.com/search?sca_esv=555561547&q="+barcode+"&tbm=shop"
    resp = requests.get(url)
    soup = BeautifulSoup(resp.text, 'lxml')
    price = soup.find_all(class_="HRLxBb")
    pricetotal = 0
    
    if len(price) > 0:
        if len(price) < 4:
            for x in range(0,len(price)):
                prices = price[x].get_text().replace("R$","").replace(",",".").strip()
                print(prices)
                pricetotal = float(prices) + pricetotal
                div = len(price)
        else:
            for x in range(0,4):
                prices = price[x].get_text().replace("R$","").replace(",",".").strip()
                pricetotal = float(prices) + pricetotal
                div = 4
    else:
        return ""
    return pricetotal/div

def get_first_price(barcode):
    url = "https://www.google.com/search?sca_esv=555561547&q="+barcode+"&tbm=shop"
    resp = requests.get(url)
    soup = BeautifulSoup(resp.text, 'lxml')
    price = soup.find(class_="HRLxBb")
    if price is None:
        return ""
    else:
        return price.get_text().replace("R$","").replace(",",".").strip()


#input = input("Entre o endereço do arquivo: ")
input = "Teste 1.xlsx"
workbook = xl.load_workbook(input)
sheet = workbook.active
i = 0
sheet.cell(row = 1, column = 2, value = "Prices")
for row in sheet.iter_rows(min_row=2, values_only=True):
    barcode = row[0]
    price = get_price(barcode)
    if price == "":
        sheet.cell(row = 2 + i, column = 2, value = price)
    else:
        sheet.cell(row = 2 + i, column = 2, value = float(price))
    i = i + 1


workbook.save("Teste 1 Concluido.xlsx")

# HRLxBb